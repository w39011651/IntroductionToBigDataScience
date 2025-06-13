import openpyxl

def save_to_xlsx(std_label, exp_label, total_data):
    try:
        workbook = openpyxl.load_workbook('player.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        workbook.create_sheet('Data')
        
    sheet = workbook.active
    sheet.delete_rows(1, sheet.max_row)
    sheet.append(['Year'] + std_label + exp_label[2:])
    
    for idx, each_year_data in enumerate(total_data):
        for player, data in each_year_data.items():
            row = [str(2003+idx),player]+data['Standard']+data['Expanded'][1:]
            sheet.append(row)

    workbook.save('playerData.xlsx')
    print("Save data done.")

"""
[outter p
    {year1
        player1: {std:data, exp:data},
        player2: {std:data, exp:data},
        .
        .
        .
    },
    {year2
        .
        .
        .
    },
    .
    .
    .
]
"""

def save_to_team_xlsx(std_label, exp_label, total_data):
    try:
        workbook = openpyxl.load_workbook('teamData.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        workbook.create_sheet('Data')
        
    sheet = workbook.active
    sheet.delete_rows(1, sheet.max_row)
    sheet.append(['YEAR'] + std_label + exp_label[2:])

    for idx, each_year_data in enumerate(total_data):
        for team, data in each_year_data.items():
            row = [str(2003+idx),team]+data['Standard']+data['Expanded'][1:]
            sheet.append(row)

    workbook.save('teamData.xlsx')
    print("Save data done.")

"""
[
    {year1
        TEAM1: {std:data, exp:data},
        TEAM2: {std:data, exp:data},
        .
        .
        .
    },
    {year2
        .
        .
        .
    },
    .
    .
    .
]
"""