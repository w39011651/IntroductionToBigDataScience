###IMPORT###
import requests
import time
from bs4 import BeautifulSoup
import openpyxl

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
###IMPORT###

###CONSTANT###
STD_POSTFIX = 'standard=True&playerPool=ALL'
EXP_POSTFIX = 'expanded=True&playerPool=ALL'
YEAR_START = 2003
YEAR_RANGE = range(2003, 2024)
MAX_PAGE_SELECTOR = '#stats-app-root > section > section > div.stats-body-table.player > div.pagination-wrapper-wvX_I3Ob > div > div > div.bui-button-group.pagination.bui-button-group > div:nth-child(7) > button > span'
PLAYER_URL = "https://www.mlb.com/stats/pitching/"
TEAM_URL = "https://www.mlb.com/stats/team/pitching/"
COOKIE_BUTTON_SELECTOR = '#onetrust-accept-btn-handler'
EXPANDED_BUTTON_SELECTOR = '#stats-app-root > section > section > div.stats-navigation > div.stats-type-wrapper-KRUAKJx8 > div > div:nth-child(1) > div > div:nth-child(2)'
PLAYER_TEAM_TABLE_SELECTOR = 'thead > tr > th > button > div > abbr'
DATA_TABLE_SELECTOR = 'thead > tr > th > button > div > span > span > abbr'
###CONSTANT###

###CONTAINER###
player_year_list = []
team_player_list = []
#player_data = dict()
team_data = dict()
std_col_label = []
exp_col_label = []
std_col_label_pt = None
exp_col_label_pt = None
driver = None
my_session = None
###CONTAINER###

###FUNCTION DEFINATION###
def log(msg, end='\n'):
    print(time.strftime('[%Y-%m-%d, %H:%M:%S]')+msg, end=end)

def print_data(container:dict):
    for key, value in container.items():
        print('NAME:'+str(key).ljust(15)+'\t'+"DATA:"+str(value['standard_data']).rjust(100))
        print('NAME:'+str(key).ljust(15)+'\t'+"DATA:"+str(value['expanded_data']).rjust(100))

def check_tag(container):
    if len(container) == 0:
        print("No tag found")
    else:
        for tag in container:
            print(tag, end = '\n\n')

def player_page_process(year, page_num):
    global player_data

    log("Processing Year{}, Page{}...".format(year, page_num))
    url = PLAYER_URL+str(year)

    if page_num != 1:
        url = url + '?page={}'.format(page_num)

    source = get_standard_and_expanded_html(url, year)
    std_data = source[0]
    exp_data = source[1]
    std_soup = BeautifulSoup(std_data, 'lxml') 

    std_players_data_table = std_soup.select('tbody > tr')
    std_each_players = [BeautifulSoup(str(players_lb), 'lxml') for players_lb in std_players_data_table]

    exp_soup = BeautifulSoup(exp_data, 'lxml')

    exp_players_data_table = exp_soup.select('tbody > tr')
    exp_each_players = [BeautifulSoup(str(players_lb), 'lxml') for players_lb in exp_players_data_table]
    player_data = player_year_list[-1]
    for p in std_each_players:
        player_name = p.select('th>div>div>div>div>a>span:nth-of-type(2)')[0].text
        player_data[player_name] = {'standard_data':[tag.text for tag in p.select('td')], 'expanded_data':None}

    for p in exp_each_players:
        player_name = p.select('th>div>div>div>div>a>span:nth-of-type(2)')[0].text
        player_data[player_name]['expanded_data'] = [tag.text for tag in p.select('td')]

    

def player_year_process(year, in_page = 1):
    global std_col_label
    global exp_col_label
    global std_col_label_pt
    global exp_col_label_pt
    try:
        log("Processing Year{}...".format(year))
        #response = requests.get(PLAYER_URL+str(year))
        retry_count = 0
        player_year_list.append(dict())
        
        response = get_standard_and_expanded_html(PLAYER_URL+str(year), 1)#1 for first page
        std_soup = BeautifulSoup(response[0], 'lxml')
        exp_soup = BeautifulSoup(response[1], 'lxml')
        year_max_page = int(std_soup.select_one(MAX_PAGE_SELECTOR).text)
        

        while len(std_col_label) == 0 or len(exp_col_label) == 0:
            log("Retry times:{}".format(retry_count))
            retry_count+=1
            response = get_standard_and_expanded_html(PLAYER_URL+str(year), 1)
            std_soup = BeautifulSoup(response[0], 'lxml')
            exp_soup = BeautifulSoup(response[1], 'lxml')
            std_col_label = [lb.text for lb in std_soup.select(DATA_TABLE_SELECTOR)]
            exp_col_label = [lb.text for lb in exp_soup.select(DATA_TABLE_SELECTOR)]

        #player_page_process(year, in_page)
        for page in range(year_max_page):
            player_page_process(year, page+1)
            print_data(player_data)

    except Exception as e:
        player_make_excel(std_col_label, exp_col_label)
        print(e)
        exit()
        
    
    

def player_make_excel(standard_column_label, expanded_column_label):
    log(str(standard_column_label))
    log(str(expanded_column_label))
    try:
        workbook = openpyxl.load_workbook('player.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        workbook.create_sheet('Data')
    try:
        for item in player_year_list:
            print(item)
        sheet = workbook.active
        sheet.delete_rows(1, sheet.max_row)
        sheet.append(['Year','PLAYER','TEAM']+standard_column_label+expanded_column_label)

        for idx, year_data in enumerate(player_year_list):
            for player, data in year_data.items():
                row = [str(2003+idx),player]+data['standard_data']+data['expanded_data'][1:]
                sheet.append(row)

        workbook.save('playerData.xlsx')
        log('Save table correctly.')
    except Exception as e:
        workbook.save('playerData.xlsx')
        log('Exception Occur, Save table done.')
        print(e)
        raise e

def get_standard_and_expanded_html(url, page):
    global driver
    STANDARD_PAGE_INDEX = 0
    EXPANDED_PAGE_INDEX = 1
    ret = [None, None]
    
    response_std = my_session.get(url+'?page={}'.format(page)+STD_POSTFIX)
    response_exp = my_session.get(url+'?page={}'.format(page)+EXP_POSTFIX)


    if response_std.status_code != 200:
        print(response_std.status_code)
        exit()
    if response_exp.status_code != 200:
        print(response_exp.status_code)
        exit()

    log('request GET over, URL={}, page={}'.format(url, page))
    
    ret[STANDARD_PAGE_INDEX] = response_std.text
    ret[EXPANDED_PAGE_INDEX] = response_exp.text

    return ret

def team_year_process(year):
    global std_col_label_pt, exp_col_label_pt, std_col_label, exp_col_label
    try:
        log("Processing Team Year{}...".format(year))
        retry_count = 0

        response = get_standard_and_expanded_html(TEAM_URL+str(year), 1)#1 for first page
        std_soup = BeautifulSoup(response[0], 'lxml')
        exp_soup = BeautifulSoup(response[1], 'lxml')
        
        std_col_label_pt = [lb.text for lb in std_soup.select(PLAYER_TEAM_TABLE_SELECTOR)]
        exp_col_label_pt = [lb.text for lb in exp_soup.select(PLAYER_TEAM_TABLE_SELECTOR)]
        std_col_label = [lb.text for lb in std_soup.select(DATA_TABLE_SELECTOR)]
        exp_col_label = [lb.text for lb in exp_soup.select(DATA_TABLE_SELECTOR)]

        while len(std_col_label) == 0 or len(exp_col_label) == 0:
            log("Retry times:{}".format(retry_count))
            retry_count+=1
            response = get_standard_and_expanded_html(TEAM_URL+str(year), 1)
            std_soup = BeautifulSoup(response[0], 'lxml')
            exp_soup = BeautifulSoup(response[1], 'lxml')
            std_col_label = [lb.text for lb in std_soup.select(DATA_TABLE_SELECTOR)]
            exp_col_label = [lb.text for lb in exp_soup.select(DATA_TABLE_SELECTOR)]

        team_page_process(year)
        print_data(team_data)
    except Exception as e:
        print("Team year process has exception, {}".format(e))
        team_make_excel(std_col_label, exp_col_label)


def team_page_process(year):
    global team_data

    log("Processing Team page, Year{}".format(year))
    url = TEAM_URL+str(year)

    source = get_standard_and_expanded_html(url, year)
    std_data = source[0]
    exp_data = source[1]
    std_soup = BeautifulSoup(std_data, 'lxml') 

    std_teams_data_table = std_soup.select('tbody > tr')
    std_each_teams = [BeautifulSoup(str(teams_lb), 'lxml') for teams_lb in std_teams_data_table]

    exp_soup = BeautifulSoup(exp_data, 'lxml')

    exp_teams_data_table = exp_soup.select('tbody > tr')
    exp_each_teams = [BeautifulSoup(str(teams_lb), 'lxml') for teams_lb in exp_teams_data_table]

    for t in std_each_teams:
        team_name = t.select('th>div>div>div>div>a>span:nth-of-type(2)')[0].text
        team_data[team_name] = {'year':year, 'standard_data':[tag.text for tag in t.select('td')], 'expanded_data':None}

    for t in exp_each_teams:
        team_name = t.select('th>div>div>div>div>a>span:nth-of-type(2)')[0].text
        team_data[team_name]['expanded_data'] = [tag.text for tag in t.select('td')]


def team_make_excel(standard_column_label, expanded_column_label):
    log(str(standard_column_label))
    log(str(expanded_column_label))
    try:
        workbook = openpyxl.load_workbook('team.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        workbook.create_sheet('Data')
    try:
        sheet = workbook.active
        sheet.delete_rows(1, sheet.max_row)
        sheet.append(['Year','TEAM','LEAGUE']+standard_column_label+expanded_column_label)

        for team, data in team_data.items():
            row = [data['year'],team]+data['standard_data']+data['expanded_data'][1:]
            sheet.append(row)

        workbook.save('teamData.xlsx')
        log('Table save correctly.')
    except:
        workbook.save('teamData.xlsx')
        log('Exception Occur, Save table done.')
    

###FUNCTION DEFINATION###

###MAIN###
if __name__ == '__main__':
    #driver = webdriver.Chrome()
    #driver.implicitly_wait(10)#set webdriver
    my_session = requests.Session()

    try:
        for year in range(2003, 2006):
            player_year_process(year)
        player_make_excel(std_col_label, exp_col_label)
    except Exception as e:
        print(e)
        exit()
    

    """for year in YEAR_RANGE:
        player_year_process(year)
    player_make_excel()

    for year in YEAR_RANGE:
        team_year_process(year)    
    
    team_make_excel(std_col_label, exp_col_label)"""

    """
    directory structure
    |
    |-webdriver.exe
    |-python file
    |-players_data.xlsx
    |-teams_data.xlsx
    """

    
    

###MAIN###